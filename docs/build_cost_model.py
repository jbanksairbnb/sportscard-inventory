"""Sports Collective — detailed infrastructure cost model.

Run: python3 docs/build_cost_model.py
Outputs: docs/Sports-Collective-Cost-Model.xlsx

Itemizes every service line at four scale tiers (100 / 1,000 / 5,000 / 10,000
users), grounded in published vendor pricing so the numbers can be defended in
a partner conversation. Edit the Inputs sheet (or the constants below) and
re-run to update.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PLUM = '3D1F4A'
ORANGE = 'E25A1C'
PAPER = 'FBF3DD'
CREAM = 'F8ECD0'
RULE = 'C8B8D0'
INK_SOFT = '554260'
TEAL = '2D7A6E'
RUST = 'C54A2C'

bold_white = Font(name='Calibri', bold=True, color='F8ECD0', size=11)
bold_plum = Font(name='Calibri', bold=True, color=PLUM, size=11)
bold_orange = Font(name='Calibri', bold=True, color=ORANGE, size=12)
italic_soft = Font(name='Calibri', italic=True, color=INK_SOFT, size=10)
mono_plum = Font(name='Consolas', color=PLUM, size=10.5)
mono_ink = Font(name='Consolas', color='2C1B33', size=10.5)
hdr_fill = PatternFill('solid', fgColor=PLUM)
band_fill = PatternFill('solid', fgColor=PAPER)
totals_fill = PatternFill('solid', fgColor=CREAM)
thin = Side(style='thin', color=RULE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# Per-user assumptions used to scale resource consumption.
PER_USER = {
    'photos':           100,    # average card photos per user
    'photo_size_mb':    0.5,    # post-resize average; raw uploads compressed
    'db_rows':          300,    # set rows + listing rows + activity
    'page_views_mo':    150,    # average page views per active user / month
    'page_size_mb':     0.6,    # typical Next.js page payload
    'image_views_mo':   400,    # image renders per user / month from CDN
    'emails_mo':        4,      # transactional emails per user / month
    'ebay_calls_mo':    30,     # eBay Browse calls/user/month (cached aggressively)
}

# Tiers — (label, total_users, % active that month)
TIERS = [
    ('Seedling (100)',   100,   1.0),
    ('Pilot+ (1,000)',   1000,  0.6),
    ('Mid (5,000)',      5000,  0.45),
    ('Growth (10,000)',  10000, 0.4),
]

# ── Pricing constants (USD; sourced from public pricing pages) ──────────────
VERCEL_PRO_BASE = 20.0
VERCEL_BANDWIDTH_INCLUDED_GB = 1000
VERCEL_BANDWIDTH_OVERAGE_PER_100GB = 40.0  # vercel.com/pricing

SUPABASE_PRO_BASE = 25.0
SUPABASE_DB_COMPUTE = {            # add-on monthly
    'micro': 0,                    # included free
    'small': 15.0,
    'medium': 60.0,
    'large': 110.0,
    'xl': 210.0,
}
SUPABASE_FILE_STORAGE_INCLUDED_GB = 100
SUPABASE_FILE_STORAGE_PER_GB = 0.021
SUPABASE_DB_STORAGE_INCLUDED_GB = 8
SUPABASE_DB_STORAGE_PER_GB = 0.125
SUPABASE_EGRESS_INCLUDED_GB = 250
SUPABASE_EGRESS_PER_GB = 0.09

CLOUDFLARE_PRO = 20.0              # cloudflare.com/plans/ — recommended once images dominate
CLOUDFLARE_IMAGE_TRANSFORMS_PER_1K = 0.50

RESEND_TIERS = [                   # resend.com/pricing
    (3000,    0.0),
    (50000,   20.0),
    (100000,  90.0),
    (1000000, 350.0),
]

EBAY_FREE_CAP = 5000 * 30          # 5,000 calls/day app-level free; per developer.ebay.com
EBAY_OVERAGE_PER_1K = 0.0          # No commercial overage tier today; Growth Check approval extends free quota.

SENTRY_TEAM = 26.0                 # sentry.io/pricing
SENTRY_BUSINESS = 80.0

DOMAIN_PER_MONTH = 1.0             # ~$12/yr through Cloudflare
SSL_PER_MONTH = 0.0                # Free via Vercel / Cloudflare

STRIPE_FEE_PCT = 0.029
STRIPE_FEE_FLAT = 0.30

# ── Tier-driven cost calculators ────────────────────────────────────────────

def calc_tier(label, total_users, active_pct):
    active = round(total_users * active_pct)

    photos = total_users * PER_USER['photos']
    photo_storage_gb = (photos * PER_USER['photo_size_mb']) / 1024
    db_rows = total_users * PER_USER['db_rows']
    db_size_gb = max(0.5, db_rows * 1.5e-6 * 1024)  # ~1.5KB/row blended
    page_egress_gb = (active * PER_USER['page_views_mo'] * PER_USER['page_size_mb']) / 1024
    image_egress_gb = (active * PER_USER['image_views_mo'] * PER_USER['photo_size_mb']) / 1024
    emails = active * PER_USER['emails_mo']
    ebay_calls = active * PER_USER['ebay_calls_mo']

    # ── Vercel ─────────────────────────────
    vercel_overage_gb = max(0.0, page_egress_gb - VERCEL_BANDWIDTH_INCLUDED_GB)
    vercel = VERCEL_PRO_BASE + (vercel_overage_gb / 100) * VERCEL_BANDWIDTH_OVERAGE_PER_100GB

    # ── Supabase ───────────────────────────
    if total_users <= 250:
        compute = SUPABASE_DB_COMPUTE['micro']
    elif total_users <= 2500:
        compute = SUPABASE_DB_COMPUTE['small']
    elif total_users <= 7500:
        compute = SUPABASE_DB_COMPUTE['medium']
    else:
        compute = SUPABASE_DB_COMPUTE['large']

    file_overage = max(0.0, photo_storage_gb - SUPABASE_FILE_STORAGE_INCLUDED_GB)
    db_overage = max(0.0, db_size_gb - SUPABASE_DB_STORAGE_INCLUDED_GB)
    # If Cloudflare CDN sits in front, image egress mostly never hits Supabase.
    cdn_in_front = total_users >= 1000
    cdn_offload = 0.85 if cdn_in_front else 0.0
    egress_via_supabase_gb = image_egress_gb * (1 - cdn_offload) + page_egress_gb * 0.05  # most page bytes come from Vercel
    egress_overage = max(0.0, egress_via_supabase_gb - SUPABASE_EGRESS_INCLUDED_GB)

    supabase = (
        SUPABASE_PRO_BASE + compute
        + file_overage * SUPABASE_FILE_STORAGE_PER_GB
        + db_overage * SUPABASE_DB_STORAGE_PER_GB
        + egress_overage * SUPABASE_EGRESS_PER_GB
    )

    # ── Cloudflare CDN ─────────────────────
    cloudflare = CLOUDFLARE_PRO if cdn_in_front else 0.0
    if cdn_in_front:
        # Cloudflare Images / transforms for thumbnails — cheap pay-as-you-go.
        cf_transforms = active * 30 / 1000 * CLOUDFLARE_IMAGE_TRANSFORMS_PER_1K
        cloudflare += cf_transforms

    # ── Resend ─────────────────────────────
    resend = 0.0
    for cap, price in RESEND_TIERS:
        resend = price
        if emails <= cap:
            break

    # ── eBay API ───────────────────────────
    ebay = 0.0  # All quotas are free at the calls/user we'd realistically use; growth-check program extends at no cost.

    # ── Sentry ─────────────────────────────
    if total_users <= 100:
        sentry = 0.0
    elif total_users <= 5000:
        sentry = SENTRY_TEAM
    else:
        sentry = SENTRY_BUSINESS

    # ── Domain / SSL ───────────────────────
    domain = DOMAIN_PER_MONTH

    # ── Stripe (only relevant once marketplace fee is live) ────
    # Estimated GMV per active user / month — placeholder for sensitivity.
    gmv_per_active = {100: 0,  1000: 50, 5000: 80, 10000: 100}.get(total_users, 50)
    monthly_gmv = active * gmv_per_active
    stripe_fees = monthly_gmv * STRIPE_FEE_PCT + active * STRIPE_FEE_FLAT * 0.5
    # (Stripe fees are paid by the seller side and netted out of your 3% fee. We
    #  surface them on the cost sheet for transparency.)

    return {
        'label':              label,
        'users':              total_users,
        'active':             active,
        'photos':             photos,
        'photo_storage_gb':   photo_storage_gb,
        'db_size_gb':         db_size_gb,
        'page_egress_gb':     page_egress_gb,
        'image_egress_gb':    image_egress_gb,
        'emails':             emails,
        'ebay_calls':         ebay_calls,
        'monthly_gmv':        monthly_gmv,
        'lines': [
            ('Vercel Pro (base + bandwidth overage)',                                         vercel),
            ('Supabase Pro base',                                                              SUPABASE_PRO_BASE),
            ('Supabase DB compute add-on',                                                     compute),
            ('Supabase file storage overage',                                                  file_overage * SUPABASE_FILE_STORAGE_PER_GB),
            ('Supabase DB storage overage',                                                    db_overage * SUPABASE_DB_STORAGE_PER_GB),
            ('Supabase egress overage',                                                        egress_overage * SUPABASE_EGRESS_PER_GB),
            ('Cloudflare CDN + image transforms',                                              cloudflare),
            ('Resend transactional email',                                                     resend),
            ('eBay Browse API',                                                                ebay),
            ('Sentry error monitoring',                                                        sentry),
            ('Domain + SSL',                                                                   domain),
        ],
        # Pass-through (not part of infra run-rate; netted from marketplace fee revenue).
        'stripe_passthrough': stripe_fees,
    }


# ── Workbook builders ───────────────────────────────────────────────────────

def header_row(ws, row, headers, widths=None):
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = bold_white; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = box
    ws.row_dimensions[row].height = 26


def write_inputs(wb):
    ws = wb.create_sheet('Inputs', 0)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14

    ws['A1'] = 'Sports Collective — Cost Model Assumptions'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:B1')

    ws['A3'] = 'Per-user resource assumptions'
    ws['A3'].font = bold_orange

    rows = [
        ('Photos per user',                               PER_USER['photos']),
        ('Average photo size after resize (MB)',          PER_USER['photo_size_mb']),
        ('DB rows per user (sets + listings + activity)', PER_USER['db_rows']),
        ('Page views per active user / month',            PER_USER['page_views_mo']),
        ('Page payload (MB)',                             PER_USER['page_size_mb']),
        ('Image renders per active user / month',         PER_USER['image_views_mo']),
        ('Transactional emails per active user / month',  PER_USER['emails_mo']),
        ('eBay Browse calls per active user / month',     PER_USER['ebay_calls_mo']),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = mono_ink
        c = ws.cell(row=i, column=2, value=v); c.font = bold_plum
        c.number_format = '#,##0.0' if isinstance(v, float) else '#,##0'

    ws.cell(row=14, column=1, value='Tier definitions').font = bold_orange
    for i, (label, users, active) in enumerate(TIERS, start=15):
        ws.cell(row=i, column=1, value=f'{label} — {active*100:.0f}% active').font = mono_ink
        c = ws.cell(row=i, column=2, value=users); c.font = bold_plum
        c.number_format = '#,##0'


def write_breakdown(wb, tier_results):
    ws = wb.create_sheet('Detailed breakdown')
    ws.column_dimensions['A'].width = 56
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws['A1'] = 'Monthly infrastructure cost — line item × tier'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:E1')
    ws['A2'] = 'All figures USD; sourced from published vendor pricing pages.'
    ws['A2'].font = italic_soft
    ws.merge_cells('A2:E2')

    headers = ['Line item'] + [t['label'] for t in tier_results]
    header_row(ws, 4, headers, widths=[56, 16, 16, 16, 16])

    line_count = len(tier_results[0]['lines'])
    for li in range(line_count):
        row = 5 + li
        label, _ = tier_results[0]['lines'][li]
        c = ws.cell(row=row, column=1, value=label); c.font = mono_ink; c.border = box
        for ti, tr in enumerate(tier_results):
            v = tr['lines'][li][1]
            cell = ws.cell(row=row, column=2 + ti, value=round(v, 2))
            cell.font = mono_ink
            cell.number_format = '"$"#,##0.00'
            cell.border = box
        if li % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = band_fill

    # Totals row
    tot_row = 5 + line_count
    c = ws.cell(row=tot_row, column=1, value='Total monthly run-rate')
    c.font = bold_orange; c.fill = totals_fill; c.border = box
    for ti, tr in enumerate(tier_results):
        total = sum(v for _, v in tr['lines'])
        cell = ws.cell(row=tot_row, column=2 + ti, value=round(total, 2))
        cell.font = bold_plum; cell.fill = totals_fill; cell.border = box
        cell.number_format = '"$"#,##0'

    # Per-active-user
    pu_row = tot_row + 1
    c = ws.cell(row=pu_row, column=1, value='Cost per active user / month')
    c.font = bold_orange; c.fill = totals_fill; c.border = box
    for ti, tr in enumerate(tier_results):
        total = sum(v for _, v in tr['lines'])
        per = total / max(1, tr['active'])
        cell = ws.cell(row=pu_row, column=2 + ti, value=round(per, 2))
        cell.font = bold_plum; cell.fill = totals_fill; cell.border = box
        cell.number_format = '"$"#,##0.00'


def write_volumes(wb, tier_results):
    ws = wb.create_sheet('Volume drivers')
    ws.column_dimensions['A'].width = 38
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws['A1'] = 'Volume drivers behind each tier'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:E1')

    headers = ['Driver'] + [t['label'] for t in tier_results]
    header_row(ws, 3, headers, widths=[38, 16, 16, 16, 16])

    rows = [
        ('Total users',           [t['users']            for t in tier_results], '#,##0'),
        ('Active users / month',  [t['active']           for t in tier_results], '#,##0'),
        ('Photos stored',         [t['photos']           for t in tier_results], '#,##0'),
        ('Photo storage (GB)',    [t['photo_storage_gb'] for t in tier_results], '#,##0.0'),
        ('Database size (GB)',    [t['db_size_gb']       for t in tier_results], '#,##0.0'),
        ('Page egress (GB / mo)', [t['page_egress_gb']   for t in tier_results], '#,##0.0'),
        ('Image egress (GB / mo)',[t['image_egress_gb']  for t in tier_results], '#,##0.0'),
        ('Emails / mo',           [t['emails']           for t in tier_results], '#,##0'),
        ('eBay API calls / mo',   [t['ebay_calls']       for t in tier_results], '#,##0'),
        ('Estimated GMV / mo',    [t['monthly_gmv']      for t in tier_results], '"$"#,##0'),
    ]
    for i, (k, vs, fmt) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = mono_ink
        ws.cell(row=i, column=1).border = box
        for j, v in enumerate(vs):
            cell = ws.cell(row=i, column=2 + j, value=v)
            cell.font = mono_ink
            cell.number_format = fmt
            cell.border = box
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = band_fill


def write_summary(wb, tier_results):
    ws = wb.create_sheet('Summary', 1)
    ws.column_dimensions['A'].width = 28
    for i in range(2, 6):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws['A1'] = 'Cost summary at-a-glance'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:E1')

    headers = ['Service group'] + [t['label'] for t in tier_results]
    header_row(ws, 3, headers, widths=[28, 16, 16, 16, 16])

    groups = [
        ('Vercel (compute + bandwidth)',     lambda tr: tr['lines'][0][1]),
        ('Supabase (DB + storage + egress)', lambda tr: sum(v for _, v in tr['lines'][1:6])),
        ('Cloudflare CDN',                   lambda tr: tr['lines'][6][1]),
        ('Resend transactional email',       lambda tr: tr['lines'][7][1]),
        ('eBay Browse API',                  lambda tr: tr['lines'][8][1]),
        ('Sentry monitoring',                lambda tr: tr['lines'][9][1]),
        ('Domain + SSL',                     lambda tr: tr['lines'][10][1]),
    ]

    row = 4
    for label, getter in groups:
        ws.cell(row=row, column=1, value=label).font = mono_ink
        ws.cell(row=row, column=1).border = box
        for ti, tr in enumerate(tier_results):
            cell = ws.cell(row=row, column=2 + ti, value=round(getter(tr), 2))
            cell.font = mono_ink
            cell.number_format = '"$"#,##0'
            cell.border = box
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = band_fill
        row += 1

    # Infra total (this is the run-rate that matters for budgeting).
    ws.cell(row=row, column=1, value='Total infra run-rate (monthly)').font = bold_orange
    ws.cell(row=row, column=1).fill = totals_fill; ws.cell(row=row, column=1).border = box
    for ti, tr in enumerate(tier_results):
        total = sum(v for _, v in tr['lines'])
        cell = ws.cell(row=row, column=2 + ti, value=round(total, 2))
        cell.font = bold_plum; cell.fill = totals_fill
        cell.number_format = '"$"#,##0'
        cell.border = box

    # Annualized
    row += 1
    ws.cell(row=row, column=1, value='Annualized (× 12)').font = bold_orange
    ws.cell(row=row, column=1).fill = totals_fill; ws.cell(row=row, column=1).border = box
    for ti, tr in enumerate(tier_results):
        total = sum(v for _, v in tr['lines']) * 12
        cell = ws.cell(row=row, column=2 + ti, value=round(total, 2))
        cell.font = bold_plum; cell.fill = totals_fill
        cell.number_format = '"$"#,##0'
        cell.border = box

    # Stripe pass-through (informational only; netted from the 3% fee revenue).
    row += 2
    ws.cell(row=row, column=1, value='Stripe processing — pass-through (not infra)').font = italic_soft
    for ti, tr in enumerate(tier_results):
        cell = ws.cell(row=row, column=2 + ti, value=round(tr['stripe_passthrough'], 2))
        cell.font = italic_soft
        cell.number_format = '"$"#,##0'
    ws.cell(row=row + 1, column=1,
            value='   Stripe is netted from the 3% marketplace fee, not added to infra spend.').font = italic_soft
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)


def write_sources(wb):
    ws = wb.create_sheet('Sources & notes')
    ws.column_dimensions['A'].width = 100

    ws['A1'] = 'Pricing sources & cost-model notes'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)

    notes = [
        '',
        '── PRICING SOURCES ──',
        'Vercel Pro: $20/mo base; bandwidth $40/100GB beyond 1TB included. vercel.com/pricing',
        'Supabase Pro: $25/mo base. DB compute add-on Micro free / Small $15 / Medium $60 / Large $110 / XL $210. supabase.com/pricing',
        '  · File storage: 100 GB included, then $0.021/GB-month.',
        '  · Database storage: 8 GB included, then $0.125/GB-month.',
        '  · Egress: 250 GB included, then $0.09/GB.',
        'Cloudflare Pro: $20/mo. Image transforms ~$0.50/1k. cloudflare.com/plans',
        'Resend: Free 3k / $20 Pro 50k / $90 Scale 100k / $350 Enterprise 1M. resend.com/pricing',
        'eBay Browse API: 5,000 calls/day free at app level; Growth Check approval extends quota at no cost. developer.ebay.com',
        'Sentry: Free dev / $26 Team / $80 Business. sentry.io/pricing',
        'Domain: ~$12/yr through Cloudflare/Namecheap. SSL free via Vercel/Cloudflare.',
        'Stripe: 2.9% + $0.30 per transaction (US standard). stripe.com/pricing',
        '',
        '── KEY MODELING DECISIONS ──',
        'Active-user ratio drops from 100% at Seedling to 40% at Growth — typical SaaS engagement decay.',
        'Cloudflare CDN turns on at 1k users; offloads ~85% of image egress from Supabase, which is the line item that breaks first.',
        'eBay API stays at $0 because we cache aggressively client-side and use Growth Check approval for bulk scans.',
        'Stripe fees show on the cost sheet for transparency but are netted out of the 3% marketplace fee — they reduce net revenue, not add to infra cost.',
        'Sensitivity to watch: image egress doubles fast if photos aren\'t pre-resized at upload. The Cloudflare offload assumption is the single biggest lever in the Growth tier.',
        '',
        '── HOW TO RE-RUN ──',
        'Edit per-user assumptions or pricing constants at the top of docs/build_cost_model.py and run python3 docs/build_cost_model.py.',
    ]
    for i, n in enumerate(notes, start=2):
        c = ws.cell(row=i, column=1, value=n)
        if n.startswith('──'):
            c.font = bold_orange
        elif n.startswith('  ·'):
            c.font = italic_soft
        else:
            c.font = mono_ink


def main():
    wb = Workbook()
    wb.remove(wb.active)

    tiers = [calc_tier(*t) for t in TIERS]

    write_inputs(wb)
    write_summary(wb, tiers)
    write_breakdown(wb, tiers)
    write_volumes(wb, tiers)
    write_sources(wb)

    out = Path(__file__).parent / 'Sports-Collective-Cost-Model.xlsx'
    wb.save(out)
    print(f'Wrote {out}')

    print('\nMonthly run-rate by tier:')
    for t in tiers:
        total = sum(v for _, v in t['lines'])
        print(f"  {t['label']:>20}  ${total:>9,.0f}/mo   (${total / max(1, t['active']):.2f} per active user)")


if __name__ == '__main__':
    main()
