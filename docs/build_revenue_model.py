"""Sports Collective — 1-year revenue model with Conservative / Base / Aggressive scenarios.

Run: python3 docs/build_revenue_model.py
Outputs: docs/Sports-Collective-Revenue-Model.xlsx

Model assumptions (editable on the Inputs sheet):
- Pricing: Collector $9/mo, Seller $24/mo, marketplace fee 3% on GMV.
- Conversion: % of free signups that upgrade. Split between Collector / Seller tiers.
- GMV per Seller: average $/mo gross merchandise sold.
- Costs scale stepwise per the cost-to-grow model (Vercel + Supabase + Resend + Sentry + domain).

Outputs three sheets — one per scenario — plus an Inputs sheet you can tweak.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PLUM = '3D1F4A'
ORANGE = 'E25A1C'
PAPER = 'FBF3DD'
CREAM = 'F8ECD0'
RULE = 'C8B8D0'
INK_SOFT = '554260'
TEAL = '2D7A6E'
RUST = 'C54A2C'

bold_white = Font(name='Calibri', bold=True, color='F8ECD0', size=11)
bold_plum = Font(name='Calibri', bold=True, color=PLUM, size=11)
bold_orange = Font(name='Calibri', bold=True, color=ORANGE, size=11)
italic_soft = Font(name='Calibri', italic=True, color=INK_SOFT, size=10)
mono_plum = Font(name='Consolas', color=PLUM, size=10.5)
mono_ink = Font(name='Consolas', color='2C1B33', size=10.5)
hdr_fill = PatternFill('solid', fgColor=PLUM)
band_fill = PatternFill('solid', fgColor=PAPER)
totals_fill = PatternFill('solid', fgColor=CREAM)
thin = Side(style='thin', color=RULE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# Cost steps — keyed by total user count (active = signups so far).
COST_STEPS = [
    (0,     46.0,   'Seedling (≤100 users)'),
    (100,   46.0,   'Seedling (≤100 users)'),
    (250,   95.0,   'Approaching pilot+ (~250 users)'),
    (500,   140.0,  'Pilot+ band (~500 users)'),
    (1000,  170.0,  'Pilot+ band (~1k users)'),
    (2500,  300.0,  'Mid-growth (~2.5k users)'),
    (5000,  600.0,  'Pre-Growth (~5k users)'),
    (10000, 1200.0, 'Growth band (~10k users)'),
]

def cost_at(users):
    """Step function — pick the highest threshold ≤ users."""
    pick = COST_STEPS[0][1]
    for thr, amt, _ in COST_STEPS:
        if users >= thr:
            pick = amt
    return pick


def add_inputs_sheet(wb, scenarios):
    ws = wb.create_sheet('Inputs', 0)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws['A1'] = 'Sports Collective — Revenue Model Assumptions'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:D1')

    ws['A3'] = 'Pricing'
    ws['A3'].font = bold_orange
    rows = [
        ('Collector tier monthly price ($)', 9),
        ('Seller tier monthly price ($)', 24),
        ('Marketplace fee on GMV (%)', 0.03),
        ('Annual discount on prepay (%)', 0.17),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = mono_ink
        c = ws.cell(row=i, column=2, value=v)
        c.font = bold_plum
        if 'price' in k.lower() and '$' in k:
            c.number_format = '"$"#,##0'
        elif '%' in k:
            c.number_format = '0.0%'

    ws['A9'] = 'Scenario inputs'
    ws['A9'].font = bold_orange
    ws['B10'] = 'Conservative'; ws['C10'] = 'Base'; ws['D10'] = 'Aggressive'
    for c in (ws['B10'], ws['C10'], ws['D10']):
        c.font = bold_white; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

    inputs_table = [
        ('New free signups / month',           [10,   25,    60]),
        ('Monthly churn (paid users)',         [0.06, 0.05,  0.04]),
        ('% of total users that upgrade to paid', [0.05, 0.12, 0.20]),
        ('Of paid: Collector tier %',          [0.85, 0.70,  0.60]),
        ('Of paid: Seller tier %',             [0.15, 0.30,  0.40]),
        ('GMV per Seller / month ($)',         [200,  450,   900]),
    ]
    for i, (k, vs) in enumerate(inputs_table, start=11):
        ws.cell(row=i, column=1, value=k).font = mono_ink
        for j, v in enumerate(vs):
            c = ws.cell(row=i, column=2 + j, value=v)
            c.font = bold_plum
            if 'GMV' in k or 'signups' in k:
                c.number_format = '"$"#,##0' if 'GMV' in k else '#,##0'
            else:
                c.number_format = '0.0%'

    ws['A19'] = 'Notes'
    ws['A19'].font = bold_orange
    notes = [
        'Pricing anchored on comp data (CollX $10, MyCardPost $9, Card Ladder $20, Cardbase $13).',
        'Marketplace fee 3% beats eBay 13.25%, COMC ~15%, Whatnot ~11%; sits between MySlabs (1–3%) and PWCC (6%).',
        'Cost steps follow the cost-to-grow model (Vercel + Supabase + Resend + Sentry + domain).',
        'Numbers are illustrative — easy to override by editing this sheet and re-running the script.',
    ]
    for i, n in enumerate(notes, start=20):
        ws.cell(row=i, column=1, value='• ' + n).font = italic_soft
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)


def build_scenario(name, signups_per_month, churn, paid_conv, collector_share, seller_share, gmv_per_seller, color, wb):
    ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 8
    for i, w in enumerate([13, 13, 13, 14, 14, 14, 16, 16, 16, 16], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws['A1'] = f'Scenario: {name}'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:K1')
    ws['A2'] = (f'Signups: {signups_per_month}/mo · Paid conv: {paid_conv*100:.0f}% '
                f'(Collector {collector_share*100:.0f}% / Seller {seller_share*100:.0f}%) · '
                f'GMV per seller: ${gmv_per_seller:,}/mo · Churn: {churn*100:.0f}%')
    ws['A2'].font = italic_soft
    ws.merge_cells('A2:K2')

    # Header row
    headers = ['Mo', 'Total users', 'Paid users', 'Collectors', 'Sellers',
               'Sub MRR', 'GMV', 'Mkt fee', 'Total MRR', 'Infra cost', 'Gross margin']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = bold_white; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = box
    ws.row_dimensions[4].height = 22

    # Project 12 months. Simple model — paid users = total users × paid_conv,
    # net of monthly paid churn (we assume churn pulls from paid pool, replaced by new conversions).
    total_users = 0
    total_sub_mrr_year = 0
    total_mkt_fee_year = 0
    total_cost_year = 0

    for m in range(1, 13):
        total_users += signups_per_month
        # paid pool grows toward steady-state paid_conv but with churn drag
        target_paid = total_users * paid_conv
        # Smooth approach — just use the target, deduct ~half a month of churn for realism
        paid_users = max(0, round(target_paid * (1 - churn / 2)))
        collectors = round(paid_users * collector_share)
        sellers = paid_users - collectors

        sub_mrr = collectors * 9 + sellers * 24
        gmv = sellers * gmv_per_seller
        mkt_fee = gmv * 0.03
        total_mrr = sub_mrr + mkt_fee

        infra = cost_at(total_users)
        margin = total_mrr - infra

        total_sub_mrr_year += sub_mrr
        total_mkt_fee_year += mkt_fee
        total_cost_year += infra

        row = 4 + m
        ws.cell(row=row, column=1, value=m).font = mono_plum
        ws.cell(row=row, column=2, value=total_users).font = mono_ink
        ws.cell(row=row, column=3, value=paid_users).font = mono_ink
        ws.cell(row=row, column=4, value=collectors).font = mono_ink
        ws.cell(row=row, column=5, value=sellers).font = mono_ink
        ws.cell(row=row, column=6, value=sub_mrr).font = mono_ink
        ws.cell(row=row, column=7, value=gmv).font = mono_ink
        ws.cell(row=row, column=8, value=mkt_fee).font = mono_ink
        ws.cell(row=row, column=9, value=total_mrr).font = bold_plum
        ws.cell(row=row, column=10, value=infra).font = mono_ink
        ws.cell(row=row, column=11, value=margin).font = (
            Font(name='Consolas', bold=True, color=TEAL if margin >= 0 else RUST, size=10.5)
        )
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = box
            if col >= 6:
                cell.number_format = '"$"#,##0'
        if m % 2 == 0:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = band_fill

    # Year totals row
    last_row = 4 + 12
    sum_row = last_row + 1
    ws.cell(row=sum_row, column=1, value='Year').font = bold_orange
    ws.cell(row=sum_row, column=2, value='').alignment = Alignment(horizontal='left')
    annual = total_sub_mrr_year + total_mkt_fee_year
    cells = {
        6: total_sub_mrr_year,
        8: total_mkt_fee_year,
        9: annual,
        10: total_cost_year,
        11: annual - total_cost_year,
    }
    for col, val in cells.items():
        c = ws.cell(row=sum_row, column=col, value=val)
        c.font = bold_plum if col != 11 else Font(
            name='Consolas', bold=True, color=TEAL if (annual - total_cost_year) >= 0 else RUST, size=11)
        c.number_format = '"$"#,##0'
        c.fill = totals_fill
        c.border = box
    for col in range(1, 12):
        ws.cell(row=sum_row, column=col).fill = totals_fill
        ws.cell(row=sum_row, column=col).border = box

    ws.cell(row=sum_row + 2, column=1,
            value=f'Annual revenue ≈ ${annual:,.0f}  ·  Infra ≈ ${total_cost_year:,.0f}  ·  '
                  f'Gross margin ≈ ${annual - total_cost_year:,.0f}  ({(1 - total_cost_year/annual)*100:.0f}% if rev>0)').font = bold_orange
    ws.merge_cells(start_row=sum_row + 2, start_column=1, end_row=sum_row + 2, end_column=11)


def build_summary(wb, scenarios):
    ws = wb.create_sheet('Summary', 1)
    ws.column_dimensions['A'].width = 32
    for i, w in enumerate([16, 16, 16], start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws['A1'] = 'Year-1 summary'
    ws['A1'].font = Font(name='Calibri', bold=True, color=PLUM, size=14)
    ws.merge_cells('A1:D1')

    headers = ['Metric', 'Conservative', 'Base', 'Aggressive']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = bold_white; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        c.border = box

    # Recompute summary numbers
    rows = []
    for sc in scenarios:
        name, signups, churn, conv, csh, ssh, gmv = sc
        total_users = 0
        sub_year = 0
        fee_year = 0
        cost_year = 0
        for m in range(1, 13):
            total_users += signups
            paid = max(0, round(total_users * conv * (1 - churn / 2)))
            collectors = round(paid * csh); sellers = paid - collectors
            sub_year += collectors * 9 + sellers * 24
            fee_year += sellers * gmv * 0.03
            cost_year += cost_at(total_users)
        rows.append((name, total_users, sub_year, fee_year, sub_year + fee_year, cost_year,
                     (sub_year + fee_year) - cost_year))

    metric_rows = [
        ('Total users by Month 12',         [r[1] for r in rows], '#,##0'),
        ('Subscription revenue',            [r[2] for r in rows], '"$"#,##0'),
        ('Marketplace fee revenue (3%)',    [r[3] for r in rows], '"$"#,##0'),
        ('Annual revenue',                  [r[4] for r in rows], '"$"#,##0'),
        ('Annual infra cost',               [r[5] for r in rows], '"$"#,##0'),
        ('Gross margin',                    [r[6] for r in rows], '"$"#,##0'),
    ]
    for i, (label, vals, fmt) in enumerate(metric_rows, start=4):
        ws.cell(row=i, column=1, value=label).font = mono_ink
        for j, v in enumerate(vals):
            c = ws.cell(row=i, column=2 + j, value=v)
            if label == 'Gross margin':
                c.font = Font(name='Consolas', bold=True, color=TEAL if v >= 0 else RUST, size=11)
            else:
                c.font = bold_plum if 'revenue' in label.lower() else mono_ink
            c.number_format = fmt
            c.border = box
            if i % 2 == 0:
                c.fill = band_fill
        ws.cell(row=i, column=1).border = box
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = band_fill

    # Footnote
    note_row = 4 + len(metric_rows) + 2
    ws.cell(row=note_row, column=1,
            value='Read across the row to compare scenarios; switch to a scenario tab to see the month-by-month build.').font = italic_soft
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    scenarios = [
        # (name, signups/mo, churn, paid conversion, collector share, seller share, GMV per seller/mo)
        ('Conservative', 10, 0.06, 0.05, 0.85, 0.15, 200),
        ('Base',         25, 0.05, 0.12, 0.70, 0.30, 450),
        ('Aggressive',   60, 0.04, 0.20, 0.60, 0.40, 900),
    ]

    add_inputs_sheet(wb, scenarios)
    build_summary(wb, scenarios)

    for sc in scenarios:
        build_scenario(*sc, color=PLUM, wb=wb)

    out = Path(__file__).parent / 'Sports-Collective-Revenue-Model.xlsx'
    wb.save(out)
    print(f'Wrote {out}')


if __name__ == '__main__':
    main()
